import csv
import os

import openpyxl
from django.http import HttpResponse
from django.utils.timezone import is_aware, make_naive

from inventory.models import TrackingUnit
from monitoring.models import Observation, ObservationPhoto, QuantityEvent


def _naive(dt):
    """Strip timezone from a datetime so openpyxl can write it to a cell."""
    if dt is not None and is_aware(dt):
        return make_naive(dt)
    return dt


# ── Tracking units ────────────────────────────────────────────────────────────

_TRACKING_UNIT_HEADERS = [
    'unit_code', 'unit_type', 'container_type',
    'crop_name', 'structured_crop',
    'accession_code', 'structured_accession',
    'batch_code', 'structured_batch',
    'quantity',
    'location_text', 'structured_location',
    'has_qr_code', 'is_active', 'archived_at', 'archive_reason',
    'created_at', 'updated_at',
]


def _tracking_unit_row(unit):
    return [
        unit.unit_code,
        unit.unit_type,
        unit.container_type,
        unit.crop_name,
        unit.crop.name if unit.crop_id else '',
        unit.accession_code,
        unit.accession.accession_code if unit.accession_id else '',
        unit.batch_code,
        unit.batch.batch_code if unit.batch_id else '',
        unit.quantity,
        unit.location_text,
        unit.display_location if unit.position_id else '',
        bool(unit.qr_code),
        unit.is_active,
        unit.archived_at,
        unit.archive_reason,
        unit.created_at,
        unit.updated_at,
    ]


def _tracking_units_queryset():
    return (
        TrackingUnit.objects
        .select_related(
            'crop',
            'accession',
            'batch',
            'position__bench__screen_house__site',
        )
        .order_by('unit_code')
    )


def export_tracking_units_csv_response():
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="tracking_units.csv"'
    writer = csv.writer(response)
    writer.writerow(_TRACKING_UNIT_HEADERS)
    for unit in _tracking_units_queryset():
        writer.writerow(_tracking_unit_row(unit))
    return response


def export_tracking_units_excel_response():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Tracking Units'
    ws.append(_TRACKING_UNIT_HEADERS)
    for unit in _tracking_units_queryset():
        ws.append(_excel_row(_tracking_unit_row(unit)))
    return _excel_response(wb, 'tracking_units.xlsx')


# ── Observations ──────────────────────────────────────────────────────────────

_OBSERVATION_HEADERS = [
    'tracking_unit_code', 'observation_type', 'corrects_observation_id',
    'observed_at', 'status', 'growth_stage', 'affected_quantity',
    'affected_zone', 'water_condition', 'pest_signs', 'disease_signs',
    'action_taken', 'notes', 'created_by', 'created_at',
]


def _observation_row(obs):
    return [
        obs.tracking_unit.unit_code,
        obs.observation_type,
        obs.corrects_observation_id or '',
        obs.observed_at,
        obs.status,
        obs.growth_stage,
        obs.affected_quantity if obs.affected_quantity is not None else '',
        obs.affected_zone,
        obs.water_condition,
        obs.pest_signs,
        obs.disease_signs,
        obs.action_taken,
        obs.notes,
        obs.created_by.username if obs.created_by else '',
        obs.created_at,
    ]


def export_observations_csv_response():
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="observations.csv"'
    writer = csv.writer(response)
    writer.writerow(_OBSERVATION_HEADERS)
    for obs in (
        Observation.objects
        .select_related('tracking_unit', 'created_by')
        .order_by('created_at')
    ):
        writer.writerow(_observation_row(obs))
    return response


def export_observations_excel_response():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Observations'
    ws.append(_OBSERVATION_HEADERS)
    for obs in (
        Observation.objects
        .select_related('tracking_unit', 'created_by')
        .order_by('created_at')
    ):
        ws.append(_excel_row(_observation_row(obs)))
    return _excel_response(wb, 'observations.xlsx')


# ── Quantity events ───────────────────────────────────────────────────────────

_QUANTITY_EVENT_HEADERS = [
    'tracking_unit_code', 'event_type', 'quantity_before',
    'quantity_change', 'quantity_after', 'reason', 'event_date', 'created_by',
]


def _quantity_event_row(event):
    return [
        event.tracking_unit.unit_code,
        event.event_type,
        event.quantity_before,
        event.quantity_change,
        event.quantity_after,
        event.reason,
        event.event_date,
        event.created_by.username if event.created_by else '',
    ]


def export_quantity_events_csv_response():
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="quantity_events.csv"'
    writer = csv.writer(response)
    writer.writerow(_QUANTITY_EVENT_HEADERS)
    for event in (
        QuantityEvent.objects
        .select_related('tracking_unit', 'created_by')
        .order_by('event_date')
    ):
        writer.writerow(_quantity_event_row(event))
    return response


def export_quantity_events_excel_response():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Quantity Events'
    ws.append(_QUANTITY_EVENT_HEADERS)
    for event in (
        QuantityEvent.objects
        .select_related('tracking_unit', 'created_by')
        .order_by('event_date')
    ):
        ws.append(_excel_row(_quantity_event_row(event)))
    return _excel_response(wb, 'quantity_events.xlsx')


# ── Photo metadata ────────────────────────────────────────────────────────────

_PHOTO_METADATA_HEADERS = [
    'tracking_unit_code', 'observation_id', 'observation_status',
    'observation_date', 'image_name', 'image_url_or_path',
    'caption', 'uploaded_at',
]


def _photo_metadata_row(photo):
    return [
        photo.observation.tracking_unit.unit_code,
        photo.observation_id,
        photo.observation.status,
        photo.observation.observed_at,
        os.path.basename(photo.image.name) if photo.image else '',
        photo.image.name if photo.image else '',
        photo.caption,
        photo.uploaded_at,
    ]


def export_photo_metadata_csv_response():
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="photo_metadata.csv"'
    writer = csv.writer(response)
    writer.writerow(_PHOTO_METADATA_HEADERS)
    for photo in (
        ObservationPhoto.objects
        .select_related('observation', 'observation__tracking_unit')
        .order_by('uploaded_at')
    ):
        writer.writerow(_photo_metadata_row(photo))
    return response


def export_photo_metadata_excel_response():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Photo Metadata'
    ws.append(_PHOTO_METADATA_HEADERS)
    for photo in (
        ObservationPhoto.objects
        .select_related('observation', 'observation__tracking_unit')
        .order_by('uploaded_at')
    ):
        ws.append(_excel_row(_photo_metadata_row(photo)))
    return _excel_response(wb, 'photo_metadata.xlsx')


# ── Internal helpers ──────────────────────────────────────────────────────────

def _excel_row(row):
    """Strip timezone info from datetime values so openpyxl can write them."""
    return [_naive(v) if hasattr(v, 'tzinfo') else v for v in row]


def _excel_response(workbook, filename):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response
